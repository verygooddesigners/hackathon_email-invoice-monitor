"""Excel document parser — extracts cell values from .xlsx files using openpyxl."""

import io
import logging
from openpyxl import load_workbook

logger = logging.getLogger("invoice_monitor")


def extract_excel_text(source) -> str:
    """Extract all cell values from an Excel workbook.

    Iterates through all sheets and formats the output as readable text
    for Claude to parse.

    Args:
        source: File path (str) or bytes object of the Excel file.

    Returns:
        Readable text representation of the workbook, or an error message string.
    """
    try:
        if isinstance(source, bytes):
            wb = load_workbook(io.BytesIO(source), data_only=True, read_only=True)
        elif isinstance(source, str):
            wb = load_workbook(source, data_only=True, read_only=True)
        else:
            return "[ERROR] Invalid input type for Excel parser."

        text_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_text = []

            for row in sheet.iter_rows():
                cell_values = []
                for cell in row:
                    if cell.value is not None:
                        cell_values.append(str(cell.value).strip())
                    else:
                        cell_values.append("")

                # Only include rows that have at least one non-empty cell
                if any(v for v in cell_values):
                    rows_text.append(" | ".join(cell_values))

            if rows_text:
                text_parts.append(f"Sheet: {sheet_name}")
                for i, row_text in enumerate(rows_text, 1):
                    text_parts.append(f"Row {i}: {row_text}")
                text_parts.append("")  # Blank line between sheets

        wb.close()
        return "\n".join(text_parts).strip() if text_parts else ""

    except Exception as e:
        logger.error(f"Error parsing Excel document: {e}")
        return f"[ERROR] Could not parse Excel document: {e}"
